import pandas as pd
import re
import time
import openpyxl
from openpyxl import load_workbook

from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# 创建 ChromeOptions 对象
options = webdriver.ChromeOptions()
# options.add_argument("--headless")  # 启用无头模式
# options.add_argument("--headless")  # 启用无头模式


# 创建 Chrome 驱动器对象，传递 options 参数

driver = webdriver.Chrome(options)
# options.add_argument("--headless")  # 启用无头模式

wait = WebDriverWait(driver, 60)

url = "https://www.bv-brc.org/view/ProteinList/?and(keyword(A),keyword(Latvia),keyword(%221506%22),keyword(%222003%22))#view_tab=proteins&defaultSort=-score"
driver.get(url)


# 输入毒株， 输出查询新毒株的url
def get_url(str):
    # 获取当前url和当前毒株名，将毒株名写入到url里面进行发请求
    keywords = str.split('/')
    # 构建新的URL
    base_url = "https://www.bv-brc.org/view/ProteinList/?and(keyword({}),keyword({}),keyword({}),keyword({}))#view_tab=proteins&defaultSort=-score"
    new_url = base_url.format(*keywords)
    return new_url


# new_url = get_url(str)
# print(new_url)  # 输出新的URL
# driver.get(new_url)
#
# time.sleep(10)


def find(str, strB):
    max_len = 0
    click_id = 0
    new_url = get_url(str)
    new_url_B = get_url(strB)
    driver.get(new_url)
    # time.sleep(6)
    element = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "dgrid-status")))
    # 获取元素的文本内容
    s = element.text
    # 如果没有参查询结果
    if "0 - 0 共 0 条结果" in s:  # 第一个名字查不出来使用第二个
        driver.get(new_url_B)
        # time.sleep(6)
        element = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "dgrid-status")))
        # 获取元素的文本内容
        s = element.text
        if "0 - 0 共 0 条结果" in s:
            return None, None
    else:  # 有查询结果先获取有多少条
        index_of_gong = s.find('共')
        if index_of_gong != -1 and index_of_gong + 1 < len(s):
            counts = s[index_of_gong + 2]
            if counts == 0:
                return None, None
            # print(f"共有{counts}条数据")

    # 开始遍历合法的数据
    index = 0
    # 获取所有div列表
    # time.sleep(4)  # 加载查询结果页面
    try:
        parent_element_xpath = '/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div'
        parent_div = wait.until(EC.presence_of_element_located((By.XPATH, parent_element_xpath)))

        # parent_div = driver.find_element(By.XPATH,
        #                                  '/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div')

        # 在父元素下查找所有的 div 元素
        list_div = parent_div.find_elements(By.TAG_NAME, 'div')
        # print(len(list_div))
    except:
        # print("no")
        time.sleep(1)
    # print("yes")

    #############
    # time.sleep(6)
    # 下面处理进行筛选
    for j in range(0, len(list_div)):
        try:
            # 判断是不是查询值
            # m = driver.find_element(by=By.XPATH,
            #                           value='/html/ body / div[1] / div[3] / div[3] / div[3]/div[2]/div/div[6]/div[2]/div/div[1]/table/tr/td[3]').text

            tx = driver.find_element(by=By.XPATH,
                                     value=f'/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div/div[{j + 1}]/table/tr/td[3]').text

            # 使用括号进行分割，获取括号中的内容
            parts = tx.split("(")
            name = parts[1].split(")")[0]

            HA = driver.find_element(by=By.XPATH,
                                     value=f'/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div/div[{j + 1}]/table/tr/td[27]').text
            # print(HA)

            # / html / body / div[1] / div[3] / div[3] / div[3] / div[2] / div / div[6] / div[2] / div / div[
            #     1] / table / tr / td[27]
            #
            # / html / body / div[1] / div[3] / div[3] / div[3] / div[2] / div / div[6] / div[2] / div / div[
            #     10] / table / tr / td[27]

            # print(name)
            # print(str)

            if name.lower() != str.lower() or HA != 'HA':  # 如果不是查询值就结束该次循环
                continue
            # 获取该条记录长度
        except:
            print("tx error!")
            return None, None
        try:
            element_xpath = f'/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div/div[{j + 1}]/table/tr/td[24]'
            length = wait.until(EC.presence_of_element_located((By.XPATH, element_xpath))).text
            # length = driver.find_element(by=By.XPATH,
            #                              value=f'/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[6]/div[2]/div/div[{j + 1}]/table/tr/td[24]').text
            # print(length)
            length = int(length)
            if length > max_len:
                max_len = length
                click_id = j
        except:
            time.sleep(1)
    # 点击最长值
    try:
        list_div[click_id].click()
    except:
        time.sleep(1)
    try:
        # 目前已经获得accession的页面,
        # a = driver.find_element(by=By.XPATH,
        #                         value='/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[5]/div[3]/div[2]/div/div/table/tbody/tr[18]/td[2]/a')
        # accession = a.text

        a = wait.until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div[3]/div[3]/div[3]/div[2]/div/div[5]/div[3]/div[2]/div/div/table/tbody/tr[18]/td[2]/a')))
        accession = a.text
        # print(accession)
    except:
        time.sleep(1)
        # return None, None

    # a.click()
    # 使用新的url进入
    base_url = 'http://www.ncbi.nlm.nih.gov/nuccore/'
    target_url = base_url + accession
    # print(target_url)
    driver.get(target_url)
    # 获取序列
    # time.sleep(6)
    try:
        element = wait.until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div[1]/form/div[1]/div[4]/div/div[5]/div[2]/div[1]/div/div/pre/span[3]')))
        sequence = element.text
    # print(sequence)
    except:
        # 保存核苷酸序列
        return None, None
    # 使用正则表达式匹配和提取 translation 的值
    pattern = r'/translation="([^"]+)"'
    matches = re.findall(pattern, sequence)
    if matches:
        concatenated_values = ''.join(matches)
        concatenated_values = concatenated_values.replace(' ', '')  # 去除空格
        concatenated_values = concatenated_values.replace('\n', '')  # 去除换行符
        # print(concatenated_values)
    else:
        print("Translation values not found.")
        return None, None
    time.sleep(1)
    return accession, concatenated_values


# find(str)

# time.sleep(10)


def write_row(sheet, str, accession, database, sequence):
    row_data = [str, accession, database, sequence]
    # 在表的最后一行添加数据
    sheet.append(row_data)
    # 保存更改

str = "A/Greece/109/2003"
database = 'bv-brc'
all_data = []
#### ========================== #####
save_path = 'result/test-ans1.xlsx'
save_name = 'Sheet1'
# 加载工作簿
save_book = load_workbook(save_path)
# 选择表
save_sheet = save_book[save_name]
#### ========================== #####

# file_path = 'cv_H3N2.xlsx'
# file_path = 'cv_H3N2.xlsx'
file_path = 'test.xlsx'
sheet_name = 'Sheet1'
# 加载工作簿
workbook = load_workbook(file_path)
# 选择表
sheet = workbook[sheet_name]
# 提取列数据并转换成列表
data_list = [cell.value for column in sheet.iter_cols(min_col=1, max_col=1) for cell in column if
             cell.value is not None]
# 使用列表推导式获取第二列的所有单元格的值
data_list_B = [cell.value for column in sheet.iter_cols(min_col=2, max_col=2) for cell in column if
               cell.value is not None]
#
def main():

    # print(data_list)
    for i in range(1, len(data_list)):
        # accession = ''
        # sequence = ''
        if (i % 10 == 0):
            print(f"已经完成{i}条数据")
        accession, sequence = find(data_list[i], data_list_B[i])
        if accession is None or sequence is None:
            accession = 'null'
            sequence = 'null'
            # continue
        # else:
        # print(accession, sequence)
        # write_row(save_sheet, data_list[i], accession, database, sequence)
        row_data = [data_list[i], accession, database, sequence]
        # row_data = [str(data_list[i]), str(accession), str(database), str(sequence)]
        print(row_data)
        # 在表的最后一行添加数据
        all_data.append(row_data)
        save_sheet.append(row_data)
        # save_book.save(save_path)

        # print("save")
        # 保存更改
    # workbook.save(save_path)
    # workbook.close()
    # 将列表转换为 Pandas DataFrame
    # df = pd.DataFrame(all_data, columns=['Data', 'Accession', 'Database', 'Sequence'])
    # 保存 DataFrame 到 Excel 文件
    # excel_path = 'test1.xlsx'
    # df.to_excel(excel_path, index=False)
    save_book.save(save_path)
    print("finish")


if __name__ == '__main__':
    main()
