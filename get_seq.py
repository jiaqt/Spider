# from telnetlib import EC
import re

from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC

from selenium import webdriver
import time
import numpy
import xlrd
from multiprocessing.dummy import Pool
import xlwt
import configparser
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from omegaconf import DictConfig, OmegaConf
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException


def login(id, pwd, wait, browser):
    browser.get("https://www.epicov.org/epi3/start")
    name = wait.until(EC.presence_of_element_located((By.ID, "elogin")))
    name.send_keys(id)  # 将id替换为您要输入的值
    word = wait.until(EC.presence_of_element_located((By.ID, "epassword")))
    word.send_keys(pwd)  # 将id替换为您要输入的值
    # 使用 WebDriverWait 来等待元素加载，并执行 click 操作
    login = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Login']")))
    login.click()
    return 1


# def find(i, workbook, sheet1, sheet2, id, pwd, browser):
def find(strain, id, pwd, browser, wait):
    browser.get("https://platform.epicov.org/epi3/frontend")
    wait = WebDriverWait(browser, 20)
    data_bases = ''
    Isolate_ID = ''
    accessions = ''
    st = ''
    total_number = 0
    # print(f'正在寻找第{i}个毒株的序列：{sheet1.cell(i, 1).value}')
    # strain = sheet1.cell(i, 1).value
    # sheet2.write(i, 1, sheet1.cell(i, 1).value)
    if '/' not in strain:
        return None, None, None, None

    tag = True
    while tag:
        try:
            div_element1 = wait.until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/form/div[5]/div/div[2]/div/div/div/div[1]/div/div/div[2]/input[1]")))
            div_element1.send_keys(id)
            # 进入查询界面
            #     print("7777")
            # wait.until(EC.visibility_of_element_located(By.XPATH,
            #                                         ("//div[@class='sys-actionbar-action-ni' and contains(text(), 'Search')]")))

            # print(div_element.text)
            # 使用 execute_script() 调用 onclick 方法, 进入查询页面
            # time.sleep(10)
            # wait.until(EC.presence_of_element_located((By.XPATH, "//div[text()='Worksets']")))
            # browser.find_element(by=By.XPATH, value=
            # "/html/body/form/div[5]/div/div[2]/div/div/div/div[1]/div/div/div[2]/input[1]").send_keys(id)

            tag = False
        except:
            time.sleep(1)
    tag = True
    # print('正在输入密码')
    while tag:
        try:
            # browser.find_element(by=By.XPATH, value=
            # '/html/body/form/div[5]/div/div[2]/div/div/div/div[1]/div/div/div[2]/input[2]').send_keys(pwd)
            div_element1 = wait.until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/form/div[5]/div/div[2]/div/div/div/div[1]/div/div/div[2]/input[2]')))
            div_element1.send_keys(pwd)
            # browser.find_element(By.ID, "epassword").send_keys(pwd)
            # login_pwd.send_keys(pwd)
            tag = False
        except:
            time.sleep(2)
    tag = True
    # print('正在点击登录')
    while tag:
        try:
            # browser.find_element(by=By.XPATH, value=
            # '/html/body/form/div[5]/div/div[2]/div/div/div/div[1]/div/div/div[2]/input[3]').click()
            browser.find_element(By.XPATH, "//input[@value='Login']").click()
            # btn.click()
            tag = False
        except:
            time.sleep(1)
    # if(1 ==login(id, pwd, wait, browser)):
    tag = True

    # wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
    # try:
    #     link_element = wait.until(EC.visibility_of_element_located((By.XPATH, "//a[text()='EpiFlu™']")))
    #     browser.execute_script(link_element.get_attribute("onclick"))
    # except:
    #     time.sleep(3)
    #     return
    # 进入指定流感病毒数据库
    # 定位到 <a> 元素

    # sequence = element.text
    while tag:
        try:
            element = wait.until(EC.presence_of_element_located((By.XPATH, "//a[text()='EpiFlu™']")))
            time.sleep(2)
            # link_element = browser.find_element(By.XPATH, "//a[text()='EpiFlu™']")
            browser.execute_script(element.get_attribute("onclick"))
            time.sleep(2)
            # wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
            tag = False
        except:
            time.sleep(1)
    tag = True
    # print("已经进入EpiFlu")
    # 进入Research
    # time.sleep(3)
    wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")

    while tag:
        try:
            div_element1 = wait.until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/form/div[5]/div/div[2]/div/div[1]/div/div/div[3]")))
            # 进入查询界面
            #     print("7777")
            # wait.until(EC.visibility_of_element_located(By.XPATH,
            #                                         ("//div[@class='sys-actionbar-action-ni' and contains(text(), 'Search')]")))

            # print(div_element.text)
            # 使用 execute_script() 调用 onclick 方法, 进入查询页面
            # time.sleep(10)
            # wait.until(EC.presence_of_element_located((By.XPATH, "//div[text()='Worksets']")))
            div_element = browser.find_element(By.XPATH,
                                               "/html/body/form/div[5]/div/div[2]/div/div[1]/div/div/div[3]")
            # print(div_element)
            # div_element.click()
            browser.execute_script(div_element.get_attribute("onclick"))
            # print("click")
            tag = False
        except:
            time.sleep(1)
    tag = True

    # 确保进入查询页面后再进行下一步操作
    # time.sleep(10)
    # print(browser.current_url)
    # 输入accession并查找
    # print('正在输入strain')
    while tag:
        try:
            input = wait.until(EC.presence_of_element_located(
                (By.XPATH,
                 '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/table[3]/tbody/tr/td[2]/div/div[1]/input')))
            # sequence = element.text
            # wait.until(EC.visibility_of_element_located((By.ID, "ce_rzfvbw_ca_entry")))
            # wait.until(EC.presence_of_element_located((By.ID, "ce_rzfvbw_ca_entry")))
            # / html / body / form / div[5] / div / div[2] / div / div[2] / div / div[1] / table[3] / tbody / tr / td[
            #     2] / div / div[1] / input

            # input = browser.find_element(By.ID, "ce_rzfvbw_ca_entry")

            # input.clear()
            input.send_keys(strain)
            input.send_keys(Keys.RETURN)
            # 使用 WebDriverWait 来等待输入框中的值与预期值匹配
            time.sleep(1)
            # 获取输入框中的值
            entered_value = input.get_attribute("value")
            if entered_value == "":
                print("null")
                return None, None, None, None
                # input.send_keys(sheet1.cell(i, 1).value)
            print("Entered value:", entered_value)
            # print(sheet1.cell(i, 1).value)
            # 33333333333333333333333333
            # # 使用CSS选择器定位按钮并点击
            # button_element = browser.find_elements(By.CSS_SELECTOR, "button.sys-form-button")
            # # 执行点击操作
            # print("0000")
            # print(button_element[1].text)
            # browser.execute_script("arguments[0].dispatchEvent(new Event('click'));", button_element[1])
            # time.sleep(1)
            # --------------
            # browser.find_element(by=By.XPATH, value=
            # '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/table[3]/tbody/tr/td[2]/div/div[1]/input').send_keys(
            #     sheet1.cell(i, 1).value)
            # --------------
            tag = False
        except:
            # tag = True
            time.sleep(1)
    # 等到成功跳转进入下一步
    # time.sleep(1)
    tag = True
    # print('正在点击HA查询,开始对输入的毒株进行查询')
    while tag:
        try:
            # # 先判断是否需要查询
            # text = wait.until(EC.presence_of_element_located(
            #     (By.XPATH,
            #      "/ html/body/form/div[5]/div/div[2]/div/div[3]/div[3]/div/div[5]/div[1]/div"))).text
            # # / html / body / form / div[5] / div / div[2] / div / div[3] / div[3] / div / div[5] / div[1] / div
            # split_parts = text.split(":")
            # # total_number = 0
            # if len(split_parts) > 1:
            #     number_text = split_parts[1].split()[0]  # 获取第一个单词，即数字部分
            #     total_number = int(number_text)
            #     if total_number > 200 or total_number == 0:
            #         return None, None, None
            # print("总数大于50")
            # else:
            #     print("未找到匹配")
            #     return
            # '/ html/body/form/div[5]/div/div[2]/div/div[2]/div/div[2]/div[1]/span/span'
            # print("开始查询")
            button_element = browser.find_elements(By.CSS_SELECTOR, "button.sys-form-button")
            # 执行点击操作
            # print("找到Search按钮")
            # print(button_element[1].text)
            browser.execute_script("arguments[0].dispatchEvent(new Event('click'));", button_element[1])
            # 加入判断语句，没有指定元素时等待一秒
            # wait.until(EC.title_contains("New Page"))
            # 等待页面加载状态变为 'complete'
            # wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
            # while True:
            #     try:
            #         # 使用 EC.presence_of_element_located() 条件，等待指定的 <h2> 元素出现
            #         wait.until(EC.presence_of_element_located(
            #             (By.XPATH, "//h2[@class='sys-dtable-caption' and @id='c_rzfvbw_ka_caption']")))
            #         break  # <h2> 元素出现，跳出循环
            #     except:
            #         # print("element not found, waiting for 1 second...")
            #         time.sleep(1)  # 等待 1 秒后继续循环
            tag = False
        except:
            tag = True
            time.sleep(1)
    tag = True
    # time.sleep(20)
    # print('正在点击查找，进入哪一个accession')
    xx1 = 0
    # 开始从结果数据里面进行筛选,开始对结果数据进行保留
    # Total先加载，直接判断Total
    element = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Total:')]")))
    # element_text = element.text
    # 判断是否满足退出条件
    if "0 isolates" in element.text:
        # print("Exiting loop, condition met:", element.text)
        return None, None, None
    if "454727 isolates" in element.text:
        # print("Exiting loop, condition met:", element.text)
        return None, None, None, None

    while tag:
        try:
            wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/div[3]/table/tbody[2]/tr')))
            list_a = browser.find_elements(by=By.XPATH, value=
            '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/div[3]/table/tbody[2]/tr')
            # list_a = browser.find_elements(by=By.XPATH, value=
            # '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/div[3]/table/tbody[2]/tr')
            # print(len(list_a))
            counts = len(list_a)
            # counts = total_number
            if (counts > 27):
                return None, None, None, None
            if counts == 0:
                return None, None, None, None
            max_len_index = 0
            max_len = 0
            for j in range(0, counts):
                try:
                    tx = browser.find_element(by=By.XPATH,
                                              value=f'/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/div[3]/table/tbody[2]/tr[{j + 1}]/td[3]').text
                except:
                    time.sleep(1)
                if strain == tx:
                    try:
                        ll = list_a[j].find_elements(by=By.XPATH, value='./td')
                        if (ll[9].text == "---"):  # 如果没有HA段跳过
                            j = j + 1
                            continue
                    except:
                        time.sleep(1)
                    if max_len < int(ll[9].text.replace(',', '')):
                        max_len = max(max_len, int(ll[9].text.replace(',', '')))
                        max_len_index = j
                else:
                    continue
            try:
                browser.find_element(by=By.XPATH,
                                     value=f'/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[1]/div[3]/table/tbody[2]/tr[{max_len_index + 1}]').click()
                time.sleep(1)
            except:
                time.sleep(1)
            tag = False
        except:
            if xx1 > 120:
                return None, None, None, None
            xx1 = xx1 + 1
            time.sleep(1)
    tag = True
    # time.sleep(10)
    xx1 = 0
    # Isolate_ID = ''  # 这个属性只有GISAID存在
    # print('正在寻找最优答案')
    while tag:
        try:
            wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
            name = browser.find_elements(by=By.TAG_NAME, value='iframe')
            browser.switch_to.frame(name[0])
            # 找到Isolate
            element = wait.until(EC.presence_of_element_located((By.XPATH,
                                                                 "/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[2]/table/tbody/tr/td[1]/table[1]/tbody/tr/td[2]/div/div[1]/div[2]")))
            # 获取元素内的文本值
            Isolate_ID = element.text
            # print("Isolate ID value:", Isolate_ID)

            tr_list = browser.find_elements(by=By.XPATH, value=
            '/html/body/form/div[5]/div/div[2]/div/div[2]/div/div[10]/table/tbody/tr')
            # time.sleep(10)
            # print(tr_list)
            tag = False
        except:
            time.sleep(1)
    tag = True
    first = True
    # col = 6
    length = 0
    # accessions = ''
    # strip = ''
    # st = ''

    for tr in tr_list:
        if first:
            first = False
            continue
        strip = tr.text.split('\n')
        if strip[0] == 'HA':
            if len(strip) >= 5:
                if int(strip[2]) > length:
                    length = int(strip[2])
                    accessions = strip[4]
                    # #print(f'accession是{answer}')
            else:
                if int(strip[2]) > length:
                    length = int(strip[2])
                    accessions = strip[3]
        else:
            continue
    # print(f'最长的accession是{answer}，长度为{length}')
    if len(strip) >= 5:
        data_bases = 'NCBI'
        # 访问最长的一条数据
        browser.get(f'https://www.ncbi.nlm.nih.gov/nuccore/{accessions}?report=GenBank')
        wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
        # time.sleep(10)
        tag = True
        while tag:
            try:
                st = browser.find_element(by=By.ID, value=f'feature_{accessions}.1_CDS_0').text.split(
                    'translation="')[1].replace('\n', '').replace("'", "").replace(' ', '')
                # print(st)
                tag = False
            except:
                time.sleep(1)
        tag = True
        # sheet2.write(i, 2, Isolate_ID)
        # sheet2.write(i, 3, answer)
        # print(f"accession是{answer}")
        # sheet2.write(i, 4, 'NCBI')
        # print("数据来源是是NCBI")

    else:
        data_bases = 'GISAID'
        for j in range(1, len(tr_list)):
            if str(length) in tr_list[j].text and 'HA' in tr_list[j].text:
                handle = browser.current_window_handle
                while tag:
                    try:
                        link_list = browser.find_elements(by=By.CLASS_NAME, value='sys-form-fi-link')
                        # time.sleep(10)
                        if len(link_list) == 2 * (len(tr_list) - 1):
                            link = link_list[2 * j - 1]
                        else:
                            link = link_list[j - 1]
                        browser.execute_script("arguments[0].click();", link)
                        # time.sleep(10)
                        tag = False
                    except:
                        time.sleep(1)
                tag = True
                # print('正在切换至跳转的页面')
                while tag:
                    try:
                        wait.until(lambda browser: browser.execute_script("return document.readyState") == "complete")
                        name = browser.find_elements(by=By.TAG_NAME, value='iframe')
                        browser.switch_to.frame(name[0])

                        s = browser.find_element(by=By.XPATH, value=
                        '/html/body/form/div[5]/div/div[2]/div/div[2]/div/table/tbody/tr[5]/td[1]/pre').text
                        # time.sleep(10)
                        check_text = browser.find_element(by=By.XPATH,
                                                          value='/html/body/form/div[5]/div/div[2]/div/div[2]/div/table/tbody/tr[2]').text.replace(
                            '\n', '')

                        tag = False
                    except:
                        time.sleep(1)
                tag = True
                split = s.split()
                string1 = ''
                for k in range(0, len(split)):
                    string1 = (string1 + split[k])
                st = ''.join([i for i in string1 if not i.isdigit()])
                # print(st)
                # sheet2.write(i, 2, Isolate_ID)
                # sheet2.write(i, 3, answer)  # accession
                # sheet2.write(i, 4, 'GISAID')
                # print('answer:', answer)
                # #print('check_text:', check_text, '\n')
                break
    # sheet2.write(i, 5, st)

    # workbook.save('tt1.xls')
    # print("finish")
    # browser.quit()
    print("ok")
    return accessions, Isolate_ID, data_bases, st


options = webdriver.ChromeOptions()
# options.add_argument("--headless")  # 启用无头模式
options.add_experimental_option("excludeSwitches", ["enable-logging"])
# 禁用缓存
# options.add_argument("--disable-cache")
# options.add_argument("--disk-cache-size=0")
# options.add_argument("--media-cache-size=0")
browser = webdriver.Chrome(options=options)
browser.implicitly_wait(20)

id = "xuej136"
# id = "JiaTao"
# pwd = "251810..aaa"
pwd = "rA1W!7QZ"
# 打开查找文件
# xls = xlrd.open_workbook_xls('血清查询_去重.xls')
# xls = xlrd.open_workbook_xls('血清查询_去重差值.xls')
# xls = xlrd.open_workbook_xls('viruspr.xls')
# # print(f'打开了{xls}文件')
# names = xls.sheet_names()
# sheet1 = xls.sheet_by_name(names[0])
# nrows = sheet1.nrows
# # 通过xlwt库创建一个新的Excel工作簿和一个新的工作表
# workbook = xlwt.Workbook(encoding='utf-8')
# sheet2 = workbook.add_sheet('test-1')

################################
save_path = 'files/virus_value.xlsx'
save_name = 'Sheet1'
save_book = load_workbook(save_path)
save_sheet = save_book[save_name]

file_path = 'files/还需要查的H3N2.xlsx'
sheet_name = 'Sheet1'
workbook = load_workbook(file_path)
sheet = workbook[sheet_name]

file_path = 'files/跳过.xlsx'
sheet_name = 'Sheet1'
workbook_skip = load_workbook(file_path)
sheet_skip = workbook[sheet_name]

data_list = [cell.value for column in sheet.iter_cols(min_col=1, max_col=1) for cell in column if
             cell.value is not None]

all_data = []


def main():
    # browser1 = browser.get("https://platform.epicov.org/epi3/frontend")
    # browser.get("https://www.epicov.org/epi3/start")
    wait = WebDriverWait(browser, 20)
    # id = "xuej136"
    # pwd = "rA1W!7QZ"
    # id = "JiaTao"
    id = "xiaoqi"
    # pwd = "251aTao"
    pwd = "B6JcFVyY"
    end = 4000
    for i in range(3000, end):
        try:
            # find(i, workbook, sheet1, sheet2, id, pwd, browser1)
            accessions, Isolate_ID, data_bases, st = find(data_list[i], id, pwd, browser, wait)

            if accessions is None or Isolate_ID is None or st is None:
                accessions = 'null'
                Isolate_ID = 'null'
                st = 'null'
        except TimeoutException:
            print("跳过")
            sheet_skip.append([data_list[i]])
            continue
        rowdata = [accessions, Isolate_ID, data_bases, st]
        # all_data.append(rowdata)
        save_sheet.append(rowdata)
        print(rowdata)
        if (i % 5 == 0):
            print(f"当前数据是{data_list[i]}")
            print(f"已经完成{i}条数据")
        save_book.save(save_path)
    print("finish")
    # workbook.save('result/血清result-1.xls')


if __name__ == '__main__':
    main()
